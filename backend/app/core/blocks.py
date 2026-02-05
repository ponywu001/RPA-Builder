"""
Block 定義與處理模組

定義所有可用的 Block 類型及其執行邏輯
"""

import asyncio
import subprocess
import httpx
import json
import re
import shutil
import os
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Callable, Type
from enum import Enum
from pathlib import Path

from .image_finder import image_finder, Position
from .desktop import desktop
from .config import settings


class BlockCategory(str, Enum):
    """Block 類別"""
    ACTION = "action"      # 動作類
    CONTROL = "control"    # 控制類
    VARIABLE = "variable"  # 變數類
    ADVANCED = "advanced"  # 進階類


@dataclass
class BlockDefinition:
    """Block 定義"""
    id: str                          # Block ID
    name: str                        # 顯示名稱
    category: BlockCategory          # 類別
    description: str                 # 說明
    params: Dict[str, Any]           # 參數定義
    has_children: bool = False       # 是否有子 blocks
    has_else: bool = False           # 是否有 else 分支
    color: str = "#5b80a5"           # 顏色


@dataclass
class BlockInstance:
    """Block 實例"""
    id: str                          # Block 類型 ID
    instance_id: str                 # 唯一實例 ID
    params: Dict[str, Any]           # 參數值
    children: List["BlockInstance"] = field(default_factory=list)
    else_children: List["BlockInstance"] = field(default_factory=list)
    
    @classmethod
    def from_dict(cls, data: dict) -> "BlockInstance":
        """從字典建立實例"""
        children = [cls.from_dict(c) for c in data.get("children", [])]
        else_children = [cls.from_dict(c) for c in data.get("else_children", [])]
        return cls(
            id=data["id"],
            instance_id=data["instance_id"],
            params=data.get("params", {}),
            children=children,
            else_children=else_children,
        )
    
    def to_dict(self) -> dict:
        """轉換為字典"""
        return {
            "id": self.id,
            "instance_id": self.instance_id,
            "params": self.params,
            "children": [c.to_dict() for c in self.children],
            "else_children": [c.to_dict() for c in self.else_children],
        }


class ExecutionContext:
    """執行上下文"""
    
    def __init__(self):
        self.variables: Dict[str, Any] = {}
        self.should_stop: bool = False
        self.should_pause: bool = False
        self.should_break: bool = False
        self.should_continue: bool = False
        self.logs: List[Dict[str, Any]] = []
        self.current_step: int = 0
        self.total_steps: int = 0
    
    def set_variable(self, name: str, value: Any) -> None:
        """設定變數"""
        self.variables[name] = value
    
    def get_variable(self, name: str, default: Any = None) -> Any:
        """取得變數"""
        return self.variables.get(name, default)
    
    def log(self, message: str, level: str = "info") -> None:
        """記錄日誌"""
        import datetime
        self.logs.append({
            "timestamp": datetime.datetime.now().isoformat(),
            "level": level,
            "message": message,
        })
    
    def resolve_value(self, value: Any) -> Any:
        """解析值（處理變數引用）"""
        if isinstance(value, str) and value.startswith("${") and value.endswith("}"):
            var_name = value[2:-1]
            return self.get_variable(var_name, value)
        return value


class BlockExecutor(ABC):
    """Block 執行器基類"""
    
    @abstractmethod
    async def execute(
        self,
        block: BlockInstance,
        context: ExecutionContext,
    ) -> Any:
        """執行 Block"""
        pass


# ==================== 動作類 Blocks ====================

class ClickImageExecutor(BlockExecutor):
    """點擊圖片"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        image_path = context.resolve_value(block.params.get("image_path"))
        timeout = block.params.get("timeout", 30)
        confidence = block.params.get("confidence", settings.default_confidence)
        offset_x = block.params.get("offset_x", 0)
        offset_y = block.params.get("offset_y", 0)
        
        context.log(f"尋找圖片: {image_path} (信心度: {confidence})")
        position = await image_finder.wait_for_image(
            image_path, timeout=timeout, confidence=confidence,
            cancel_check=lambda: context.should_stop
        )
        
        if context.should_stop:
            return {"cancelled": True}
        
        if position:
            click_x = position.x + offset_x
            click_y = position.y + offset_y
            context.log(f"找到圖片 (匹配度: {position.confidence:.2%})，點擊座標 ({click_x}, {click_y})")
            desktop.click(click_x, click_y)
            return position.to_dict()
        else:
            raise TimeoutError(f"找不到圖片: {image_path}")


class ClickPositionExecutor(BlockExecutor):
    """點擊座標"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        x = int(context.resolve_value(block.params.get("x", 0)))
        y = int(context.resolve_value(block.params.get("y", 0)))
        button = block.params.get("button", "left")
        
        context.log(f"點擊座標 ({x}, {y}), 按鈕: {button}")
        desktop.click(x, y, button=button)
        return {"x": x, "y": y}


class DoubleClickImageExecutor(BlockExecutor):
    """雙擊圖片"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        image_path = context.resolve_value(block.params.get("image_path"))
        timeout = block.params.get("timeout", 30)
        confidence = block.params.get("confidence", settings.default_confidence)
        
        context.log(f"尋找圖片並雙擊: {image_path} (信心度: {confidence})")
        position = await image_finder.wait_for_image(
            image_path, timeout=timeout, confidence=confidence,
            cancel_check=lambda: context.should_stop
        )
        
        if context.should_stop:
            return {"cancelled": True}
        
        if position:
            context.log(f"找到圖片 (匹配度: {position.confidence:.2%})，雙擊座標 ({position.x}, {position.y})")
            desktop.double_click(position.x, position.y)
            return position.to_dict()
        else:
            raise TimeoutError(f"找不到圖片: {image_path}")


class RightClickImageExecutor(BlockExecutor):
    """右鍵點擊圖片"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        image_path = context.resolve_value(block.params.get("image_path"))
        timeout = block.params.get("timeout", 30)
        confidence = block.params.get("confidence", settings.default_confidence)
        
        context.log(f"尋找圖片並右鍵點擊: {image_path} (信心度: {confidence})")
        position = await image_finder.wait_for_image(
            image_path, timeout=timeout, confidence=confidence,
            cancel_check=lambda: context.should_stop
        )
        
        if context.should_stop:
            return {"cancelled": True}
        
        if position:
            context.log(f"找到圖片 (匹配度: {position.confidence:.2%})，右鍵點擊座標 ({position.x}, {position.y})")
            desktop.right_click(position.x, position.y)
            return position.to_dict()
        else:
            raise TimeoutError(f"找不到圖片: {image_path}")


class TypeTextExecutor(BlockExecutor):
    """輸入文字"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        text = str(context.resolve_value(block.params.get("text", "")))
        interval = block.params.get("interval", 0.05)
        
        context.log(f"輸入文字: {text[:50]}{'...' if len(text) > 50 else ''}")
        desktop.write(text)
        return {"text": text}


class HotkeyExecutor(BlockExecutor):
    """快捷鍵"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        keys = block.params.get("keys", [])
        if isinstance(keys, str):
            keys = [k.strip() for k in keys.split("+")]
        
        context.log(f"執行快捷鍵: {'+'.join(keys)}")
        desktop.hotkey(*keys)
        return {"keys": keys}


class ScrollExecutor(BlockExecutor):
    """滾動"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        direction = block.params.get("direction", "down")
        amount = int(block.params.get("amount", 3))
        x = block.params.get("x")
        y = block.params.get("y")
        
        # 方向處理
        clicks = amount if direction == "up" else -amount
        
        context.log(f"滾動 {direction} {amount} 格")
        
        if x is not None and y is not None:
            desktop.scroll(clicks, int(x), int(y))
        else:
            desktop.scroll(clicks)
        
        return {"direction": direction, "amount": amount}


class DragDropExecutor(BlockExecutor):
    """拖放"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        from_image = context.resolve_value(block.params.get("from_image"))
        to_image = context.resolve_value(block.params.get("to_image"))
        timeout = block.params.get("timeout", 30)
        confidence = block.params.get("confidence", settings.default_confidence)
        
        context.log(f"拖放: {from_image} -> {to_image} (信心度: {confidence})")
        
        # 找到起點
        from_pos = await image_finder.wait_for_image(
            from_image, timeout=timeout, confidence=confidence,
            cancel_check=lambda: context.should_stop
        )
        if context.should_stop:
            return {"cancelled": True}
        if not from_pos:
            raise TimeoutError(f"找不到起點圖片: {from_image}")
        
        # 找到終點
        to_pos = await image_finder.wait_for_image(
            to_image, timeout=timeout, confidence=confidence,
            cancel_check=lambda: context.should_stop
        )
        if context.should_stop:
            return {"cancelled": True}
        if not to_pos:
            raise TimeoutError(f"找不到終點圖片: {to_image}")
        
        context.log(f"從 ({from_pos.x}, {from_pos.y}) 拖到 ({to_pos.x}, {to_pos.y})")
        desktop.drag_to(from_pos.x, from_pos.y, to_pos.x, to_pos.y)
        
        return {
            "from": from_pos.to_dict(),
            "to": to_pos.to_dict(),
        }


class WaitExecutor(BlockExecutor):
    """等待"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        seconds = float(block.params.get("seconds", 1))
        
        context.log(f"等待 {seconds} 秒")
        
        # 分段等待以便能夠響應停止指令
        interval = 0.1  # 每 0.1 秒檢查一次
        elapsed = 0.0
        while elapsed < seconds:
            if context.should_stop:
                return {"cancelled": True}
            await asyncio.sleep(min(interval, seconds - elapsed))
            elapsed += interval
        
        return {"seconds": seconds}


class WaitImageExecutor(BlockExecutor):
    """等待圖片出現"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        image_path = context.resolve_value(block.params.get("image_path"))
        timeout = block.params.get("timeout", 30)
        confidence = block.params.get("confidence", settings.default_confidence)
        
        context.log(f"等待圖片出現: {image_path} (信心度: {confidence})")
        position = await image_finder.wait_for_image(
            image_path, timeout=timeout, confidence=confidence,
            cancel_check=lambda: context.should_stop
        )
        
        if context.should_stop:
            return {"cancelled": True}
        
        if position:
            context.log(f"圖片已出現於 ({position.x}, {position.y})，匹配度: {position.confidence:.2%}")
            return position.to_dict()
        else:
            raise TimeoutError(f"等待圖片超時: {image_path}")


class WaitImageGoneExecutor(BlockExecutor):
    """等待圖片消失"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        image_path = context.resolve_value(block.params.get("image_path"))
        timeout = block.params.get("timeout", 30)
        confidence = block.params.get("confidence", settings.default_confidence)
        
        context.log(f"等待圖片消失: {image_path} (信心度: {confidence})")
        result = await image_finder.wait_until_disappear(
            image_path, timeout=timeout, confidence=confidence,
            cancel_check=lambda: context.should_stop
        )
        
        if context.should_stop:
            return {"cancelled": True}
        
        if result:
            context.log("圖片已消失")
            return {"disappeared": True}
        else:
            raise TimeoutError(f"等待圖片消失超時: {image_path}")


# ==================== 控制類 Blocks ====================

class IfImageExistsExecutor(BlockExecutor):
    """如果圖片存在"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        image_path = context.resolve_value(block.params.get("image_path"))
        confidence = block.params.get("confidence", 0.8)
        
        context.log(f"檢查圖片是否存在: {image_path}")
        position = image_finder.find_on_screen(image_path, confidence=confidence)
        
        if position:
            context.log("圖片存在，執行子區塊")
            return {"exists": True, "position": position.to_dict(), "execute_children": True}
        else:
            context.log("圖片不存在，執行 else 區塊")
            return {"exists": False, "execute_else": True}


class LoopTimesExecutor(BlockExecutor):
    """重複 N 次"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        times = int(block.params.get("times", 1))
        context.log(f"開始迴圈，次數: {times}")
        return {"loop": True, "times": times}


class LoopWhileImageExecutor(BlockExecutor):
    """當圖片存在時重複"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        image_path = context.resolve_value(block.params.get("image_path"))
        max_iterations = block.params.get("max_iterations", 100)
        confidence = block.params.get("confidence", settings.default_confidence)
        
        context.log(f"當圖片存在時重複: {image_path} (信心度: {confidence})")
        return {
            "loop": True,
            "condition": "while_image",
            "image_path": image_path,
            "max_iterations": max_iterations,
            "confidence": confidence,
        }


class LoopUntilImageExecutor(BlockExecutor):
    """重複直到圖片出現"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        image_path = context.resolve_value(block.params.get("image_path"))
        max_iterations = block.params.get("max_iterations", 100)
        confidence = block.params.get("confidence", settings.default_confidence)
        
        context.log(f"重複直到圖片出現: {image_path} (信心度: {confidence})")
        return {
            "loop": True,
            "condition": "until_image",
            "image_path": image_path,
            "max_iterations": max_iterations,
            "confidence": confidence,
        }


class BreakExecutor(BlockExecutor):
    """跳出迴圈"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        context.log("跳出迴圈")
        context.should_break = True
        return {"break": True}


class ContinueExecutor(BlockExecutor):
    """繼續下一輪"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        context.log("繼續下一輪迴圈")
        context.should_continue = True
        return {"continue": True}


# ==================== 變數類 Blocks ====================

class SetVariableExecutor(BlockExecutor):
    """設定變數"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        name = block.params.get("name", "")
        value = context.resolve_value(block.params.get("value"))
        
        context.log(f"設定變數 {name} = {value}")
        context.set_variable(name, value)
        return {"name": name, "value": value}


class GetVariableExecutor(BlockExecutor):
    """取得變數"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        name = block.params.get("name", "")
        value = context.get_variable(name)
        
        context.log(f"取得變數 {name} = {value}")
        return {"name": name, "value": value}


class SavePositionExecutor(BlockExecutor):
    """儲存座標"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        variable_name = block.params.get("variable_name", "")
        image_path = context.resolve_value(block.params.get("image_path"))
        timeout = block.params.get("timeout", 30)
        confidence = block.params.get("confidence", settings.default_confidence)
        
        context.log(f"尋找圖片並儲存座標到變數 {variable_name} (信心度: {confidence})")
        position = await image_finder.wait_for_image(
            image_path, timeout=timeout, confidence=confidence,
            cancel_check=lambda: context.should_stop
        )
        
        if context.should_stop:
            return {"cancelled": True}
        
        if position:
            context.set_variable(variable_name, position.to_dict())
            context.log(f"座標已儲存: ({position.x}, {position.y})，匹配度: {position.confidence:.2%}")
            return position.to_dict()
        else:
            raise TimeoutError(f"找不到圖片: {image_path}")


# ==================== 進階類 Blocks ====================

class RunScriptExecutor(BlockExecutor):
    """執行子腳本"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        script_id = block.params.get("script_id", "")
        context.log(f"執行子腳本: {script_id}")
        # 實際執行會在 engine.py 中處理
        return {"script_id": script_id, "run_subscript": True}


class HttpRequestExecutor(BlockExecutor):
    """HTTP 請求"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        method = block.params.get("method", "GET")
        url = context.resolve_value(block.params.get("url", ""))
        headers = block.params.get("headers", {})
        body = context.resolve_value(block.params.get("body", ""))
        
        context.log(f"發送 {method} 請求到 {url}")
        
        async with httpx.AsyncClient() as client:
            response = await client.request(
                method=method,
                url=url,
                headers=headers,
                content=body if body else None,
            )
            
            context.log(f"回應狀態碼: {response.status_code}")
            
            return {
                "status_code": response.status_code,
                "headers": dict(response.headers),
                "body": response.text,
            }


class RunCommandExecutor(BlockExecutor):
    """執行命令"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        command = context.resolve_value(block.params.get("command", ""))
        
        context.log(f"執行命令: {command}")
        
        result = subprocess.run(
            command,
            shell=True,
            capture_output=True,
            text=True,
        )
        
        context.log(f"命令執行完成，返回碼: {result.returncode}")
        
        return {
            "returncode": result.returncode,
            "stdout": result.stdout,
            "stderr": result.stderr,
        }


class LogExecutor(BlockExecutor):
    """記錄日誌"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        message = str(context.resolve_value(block.params.get("message", "")))
        level = block.params.get("level", "info")
        
        context.log(message, level=level)
        return {"message": message, "level": level}


class ScreenshotExecutor(BlockExecutor):
    """截圖"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        save_path = context.resolve_value(block.params.get("save_path", ""))
        
        context.log(f"擷取螢幕截圖")
        
        screenshot_bytes = image_finder.capture_screen()
        
        if save_path:
            path = Path(save_path)
            if not path.is_absolute():
                path = settings.images_dir / path
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(screenshot_bytes)
            context.log(f"截圖已儲存到: {path}")
        
        return {"saved": bool(save_path), "path": str(save_path) if save_path else None}


class SendEmailExecutor(BlockExecutor):
    """發送 Email"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        
        to_addr = context.resolve_value(block.params.get("to", ""))
        subject = context.resolve_value(block.params.get("subject", ""))
        body = context.resolve_value(block.params.get("body", ""))
        
        # 從環境變數或配置取得 SMTP 設定
        smtp_host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
        smtp_port = int(os.environ.get("SMTP_PORT", "587"))
        smtp_user = os.environ.get("SMTP_USER", "")
        smtp_pass = os.environ.get("SMTP_PASS", "")
        
        if not smtp_user or not smtp_pass:
            context.log("警告: 未設定 SMTP 帳號密碼，請設定環境變數 SMTP_USER 和 SMTP_PASS")
            return {"sent": False, "error": "SMTP credentials not configured"}
        
        context.log(f"發送郵件到: {to_addr}")
        
        try:
            msg = MIMEMultipart()
            msg['From'] = smtp_user
            msg['To'] = to_addr
            msg['Subject'] = subject
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.send_message(msg)
            
            context.log("郵件發送成功")
            return {"sent": True, "to": to_addr}
        except Exception as e:
            context.log(f"郵件發送失敗: {e}")
            return {"sent": False, "error": str(e)}


class RetryExecutor(BlockExecutor):
    """重試執行"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        times = int(block.params.get("times", 3))
        interval = float(block.params.get("interval", 1))
        
        context.log(f"重試區塊: 最多 {times} 次，間隔 {interval} 秒")
        
        return {
            "retry": True,
            "max_times": times,
            "interval": interval,
            "current_attempt": 0,
        }


# ==================== 檔案操作 Blocks ====================

class ReadFileExecutor(BlockExecutor):
    """讀取檔案"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        file_path = context.resolve_value(block.params.get("file_path", ""))
        encoding = block.params.get("encoding", "utf-8")
        
        context.log(f"讀取檔案: {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"檔案不存在: {file_path}")
        
        content = path.read_text(encoding=encoding)
        context.log(f"讀取完成，{len(content)} 字元")
        
        return {"content": content, "path": str(path)}


class WriteFileExecutor(BlockExecutor):
    """寫入檔案"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        file_path = context.resolve_value(block.params.get("file_path", ""))
        content = str(context.resolve_value(block.params.get("content", "")))
        encoding = block.params.get("encoding", "utf-8")
        append = block.params.get("append", False)
        
        context.log(f"{'附加' if append else '寫入'}檔案: {file_path}")
        
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        
        if append:
            with open(path, 'a', encoding=encoding) as f:
                f.write(content)
        else:
            path.write_text(content, encoding=encoding)
        
        context.log(f"寫入完成，{len(content)} 字元")
        
        return {"path": str(path), "length": len(content)}


class CopyFileExecutor(BlockExecutor):
    """複製檔案"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        source = context.resolve_value(block.params.get("source", ""))
        destination = context.resolve_value(block.params.get("destination", ""))
        
        context.log(f"複製: {source} -> {destination}")
        
        src_path = Path(source)
        dst_path = Path(destination)
        
        if not src_path.exists():
            raise FileNotFoundError(f"來源不存在: {source}")
        
        dst_path.parent.mkdir(parents=True, exist_ok=True)
        
        if src_path.is_dir():
            shutil.copytree(src_path, dst_path)
        else:
            shutil.copy2(src_path, dst_path)
        
        context.log("複製完成")
        
        return {"source": str(src_path), "destination": str(dst_path)}


class MoveFileExecutor(BlockExecutor):
    """移動檔案"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        source = context.resolve_value(block.params.get("source", ""))
        destination = context.resolve_value(block.params.get("destination", ""))
        
        context.log(f"移動: {source} -> {destination}")
        
        src_path = Path(source)
        dst_path = Path(destination)
        
        if not src_path.exists():
            raise FileNotFoundError(f"來源不存在: {source}")
        
        dst_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(src_path), str(dst_path))
        
        context.log("移動完成")
        
        return {"source": str(src_path), "destination": str(dst_path)}


class DeleteFileExecutor(BlockExecutor):
    """刪除檔案"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        file_path = context.resolve_value(block.params.get("file_path", ""))
        
        context.log(f"刪除: {file_path}")
        
        path = Path(file_path)
        
        if not path.exists():
            context.log("檔案不存在，跳過")
            return {"deleted": False, "path": str(path)}
        
        if path.is_dir():
            shutil.rmtree(path)
        else:
            path.unlink()
        
        context.log("刪除完成")
        
        return {"deleted": True, "path": str(path)}


class FileExistsExecutor(BlockExecutor):
    """檢查檔案是否存在"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        file_path = context.resolve_value(block.params.get("file_path", ""))
        
        path = Path(file_path)
        exists = path.exists()
        
        context.log(f"檢查檔案: {file_path} - {'存在' if exists else '不存在'}")
        
        return {
            "exists": exists,
            "is_file": path.is_file() if exists else False,
            "is_dir": path.is_dir() if exists else False,
            "execute_children": exists,
            "execute_else": not exists,
        }


class ListFilesExecutor(BlockExecutor):
    """列出檔案"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        directory = context.resolve_value(block.params.get("directory", ""))
        pattern = block.params.get("pattern", "*")
        
        context.log(f"列出檔案: {directory} (模式: {pattern})")
        
        path = Path(directory)
        if not path.exists():
            raise FileNotFoundError(f"目錄不存在: {directory}")
        
        files = list(path.glob(pattern))
        file_list = [str(f) for f in files]
        
        context.log(f"找到 {len(file_list)} 個檔案")
        
        return {"files": file_list, "count": len(file_list)}


# ==================== Excel 操作 Blocks ====================

class ExcelReadExecutor(BlockExecutor):
    """讀取 Excel"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import openpyxl
        
        file_path = context.resolve_value(block.params.get("file_path", ""))
        sheet_name = context.resolve_value(block.params.get("sheet", "Sheet1"))
        var_name = block.params.get("var_name", "data")
        
        context.log(f"讀取 Excel: {file_path}, 工作表: {sheet_name}")
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        
        context.set_variable(var_name, data)
        context.log(f"讀取完成，{len(data)} 行資料")
        
        return {"rows": len(data), "var_name": var_name}


class ExcelWriteExecutor(BlockExecutor):
    """寫入 Excel"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import openpyxl
        
        file_path = context.resolve_value(block.params.get("file_path", ""))
        sheet_name = context.resolve_value(block.params.get("sheet", "Sheet1"))
        data_var = block.params.get("data_var", "data")
        
        data = context.get_variable(data_var)
        if not isinstance(data, list):
            raise ValueError(f"變數 {data_var} 不是列表格式")
        
        context.log(f"寫入 Excel: {file_path}, 工作表: {sheet_name}")
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        
        for row_idx, row in enumerate(data, 1):
            if isinstance(row, list):
                for col_idx, value in enumerate(row, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
        
        workbook.save(file_path)
        context.log(f"寫入完成，{len(data)} 行資料")
        
        return {"rows": len(data), "path": file_path}


class ExcelReadCellExecutor(BlockExecutor):
    """讀取 Excel 儲存格"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import openpyxl
        
        file_path = context.resolve_value(block.params.get("file_path", ""))
        sheet_name = context.resolve_value(block.params.get("sheet", "Sheet1"))
        cell = context.resolve_value(block.params.get("cell", "A1"))
        var_name = block.params.get("var_name", "value")
        
        context.log(f"讀取 Excel 儲存格: {file_path}!{sheet_name}!{cell}")
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        
        value = sheet[cell].value
        context.set_variable(var_name, value)
        context.log(f"儲存格值: {value}")
        
        return {"value": value, "var_name": var_name}


class ExcelWriteCellExecutor(BlockExecutor):
    """寫入 Excel 儲存格"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import openpyxl
        
        file_path = context.resolve_value(block.params.get("file_path", ""))
        sheet_name = context.resolve_value(block.params.get("sheet", "Sheet1"))
        cell = context.resolve_value(block.params.get("cell", "A1"))
        value = context.resolve_value(block.params.get("value", ""))
        
        context.log(f"寫入 Excel 儲存格: {file_path}!{sheet_name}!{cell} = {value}")
        
        # 如果檔案存在則載入，否則創建新的
        path = Path(file_path)
        if path.exists():
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = openpyxl.Workbook()
        
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
        
        sheet[cell] = value
        workbook.save(file_path)
        
        context.log(f"寫入完成")
        
        return {"cell": cell, "value": value}


# ==================== 錯誤處理 Blocks ====================

class TryCatchExecutor(BlockExecutor):
    """Try-Catch 錯誤處理"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        context.log("進入 Try 區塊")
        return {"try_catch": True, "execute_try": True}


# ==================== 迴圈增強 Blocks ====================

class ForEachExecutor(BlockExecutor):
    """For Each 遍歷"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        list_value = context.resolve_value(block.params.get("list", ""))
        variable_name = block.params.get("variable_name", "item")
        
        # 解析列表
        if isinstance(list_value, str):
            try:
                list_value = json.loads(list_value)
            except json.JSONDecodeError:
                # 嘗試以逗號分隔
                list_value = [x.strip() for x in list_value.split(",")]
        
        if not isinstance(list_value, list):
            list_value = [list_value]
        
        context.log(f"遍歷列表，共 {len(list_value)} 項")
        
        return {
            "loop": True,
            "foreach": True,
            "list": list_value,
            "variable_name": variable_name,
        }


# ==================== 字串處理 Blocks ====================

class StringConcatExecutor(BlockExecutor):
    """字串拼接"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        str1 = str(context.resolve_value(block.params.get("string1", "")))
        str2 = str(context.resolve_value(block.params.get("string2", "")))
        separator = block.params.get("separator", "")
        
        result = str1 + separator + str2
        context.log(f"字串拼接結果: {result[:50]}{'...' if len(result) > 50 else ''}")
        
        return {"result": result}


class StringSplitExecutor(BlockExecutor):
    """字串分割"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        text = str(context.resolve_value(block.params.get("text", "")))
        separator = block.params.get("separator", ",")
        
        parts = text.split(separator)
        context.log(f"字串分割為 {len(parts)} 個部分")
        
        return {"parts": parts, "count": len(parts)}


class StringReplaceExecutor(BlockExecutor):
    """字串替換"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        text = str(context.resolve_value(block.params.get("text", "")))
        search = str(context.resolve_value(block.params.get("search", "")))
        replace = str(context.resolve_value(block.params.get("replace", "")))
        use_regex = block.params.get("use_regex", False)
        
        if use_regex:
            result = re.sub(search, replace, text)
        else:
            result = text.replace(search, replace)
        
        context.log(f"字串替換完成")
        
        return {"result": result}


class StringMatchExecutor(BlockExecutor):
    """正則匹配"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        text = str(context.resolve_value(block.params.get("text", "")))
        pattern = str(context.resolve_value(block.params.get("pattern", "")))
        
        matches = re.findall(pattern, text)
        context.log(f"正則匹配找到 {len(matches)} 個結果")
        
        return {"matches": matches, "count": len(matches), "found": len(matches) > 0}


class StringFormatExecutor(BlockExecutor):
    """字串格式化"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        template = str(context.resolve_value(block.params.get("template", "")))
        
        # 替換模板中的變數引用
        result = template
        for key, value in context.variables.items():
            result = result.replace(f"${{{key}}}", str(value))
        
        context.log(f"格式化結果: {result[:50]}{'...' if len(result) > 50 else ''}")
        
        return {"result": result}


# ==================== 數學運算 Blocks ====================

class MathOperationExecutor(BlockExecutor):
    """數學運算"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        a = float(context.resolve_value(block.params.get("a", 0)))
        b = float(context.resolve_value(block.params.get("b", 0)))
        operation = block.params.get("operation", "add")
        
        if operation == "add":
            result = a + b
        elif operation == "subtract":
            result = a - b
        elif operation == "multiply":
            result = a * b
        elif operation == "divide":
            if b == 0:
                raise ValueError("除以零錯誤")
            result = a / b
        elif operation == "modulo":
            result = a % b
        elif operation == "power":
            result = a ** b
        else:
            raise ValueError(f"未知運算: {operation}")
        
        context.log(f"運算: {a} {operation} {b} = {result}")
        
        return {"result": result}


class RandomNumberExecutor(BlockExecutor):
    """隨機數"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import random
        
        min_val = float(context.resolve_value(block.params.get("min", 0)))
        max_val = float(context.resolve_value(block.params.get("max", 100)))
        is_integer = block.params.get("integer", True)
        
        if is_integer:
            result = random.randint(int(min_val), int(max_val))
        else:
            result = random.uniform(min_val, max_val)
        
        context.log(f"隨機數: {result}")
        
        return {"result": result}


# ==================== 剪貼簿/JSON Blocks ====================

class ClipboardReadExecutor(BlockExecutor):
    """讀取剪貼簿"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import pyperclip
        
        content = pyperclip.paste()
        context.log(f"讀取剪貼簿: {content[:50]}{'...' if len(content) > 50 else ''}")
        
        return {"content": content}


class ClipboardWriteExecutor(BlockExecutor):
    """寫入剪貼簿"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import pyperclip
        
        content = str(context.resolve_value(block.params.get("content", "")))
        pyperclip.copy(content)
        
        context.log(f"寫入剪貼簿: {content[:50]}{'...' if len(content) > 50 else ''}")
        
        return {"content": content}


class JsonParseExecutor(BlockExecutor):
    """解析 JSON"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        text = str(context.resolve_value(block.params.get("text", "")))
        
        try:
            data = json.loads(text)
            context.log("JSON 解析成功")
            return {"data": data, "success": True}
        except json.JSONDecodeError as e:
            context.log(f"JSON 解析失敗: {e}", level="error")
            return {"data": None, "success": False, "error": str(e)}


class JsonStringifyExecutor(BlockExecutor):
    """轉換為 JSON"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        data = context.resolve_value(block.params.get("data", {}))
        pretty = block.params.get("pretty", False)
        
        if pretty:
            result = json.dumps(data, ensure_ascii=False, indent=2)
        else:
            result = json.dumps(data, ensure_ascii=False)
        
        context.log(f"JSON 轉換完成")
        
        return {"result": result}


class JsonGetValueExecutor(BlockExecutor):
    """取得 JSON 值"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        data = context.resolve_value(block.params.get("data", {}))
        path = str(context.resolve_value(block.params.get("path", "")))
        
        # 支援點號路徑 (如 user.name)
        value = data
        for key in path.split("."):
            if isinstance(value, dict):
                value = value.get(key)
            elif isinstance(value, list):
                try:
                    value = value[int(key)]
                except (ValueError, IndexError):
                    value = None
                    break
            else:
                value = None
                break
        
        context.log(f"取得 JSON 值: {path} = {value}")
        
        return {"value": value}


# ==================== 視窗控制 Blocks ====================

class GetWindowExecutor(BlockExecutor):
    """取得視窗"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import pyautogui
        
        title = context.resolve_value(block.params.get("title", ""))
        
        # 使用 pyautogui 的 getWindowsWithTitle
        try:
            windows = pyautogui.getWindowsWithTitle(title)
            if windows:
                window = windows[0]
                context.log(f"找到視窗: {window.title}")
                return {
                    "found": True,
                    "title": window.title,
                    "left": window.left,
                    "top": window.top,
                    "width": window.width,
                    "height": window.height,
                }
            else:
                context.log(f"未找到視窗: {title}")
                return {"found": False}
        except Exception as e:
            context.log(f"取得視窗失敗: {e}", level="warning")
            return {"found": False, "error": str(e)}


class ActivateWindowExecutor(BlockExecutor):
    """啟用視窗"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import pyautogui
        
        title = context.resolve_value(block.params.get("title", ""))
        
        try:
            windows = pyautogui.getWindowsWithTitle(title)
            if windows:
                window = windows[0]
                window.activate()
                context.log(f"已啟用視窗: {window.title}")
                return {"activated": True, "title": window.title}
            else:
                context.log(f"未找到視窗: {title}")
                return {"activated": False}
        except Exception as e:
            context.log(f"啟用視窗失敗: {e}", level="warning")
            return {"activated": False, "error": str(e)}


class CloseWindowExecutor(BlockExecutor):
    """關閉視窗"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import pyautogui
        
        title = context.resolve_value(block.params.get("title", ""))
        
        try:
            windows = pyautogui.getWindowsWithTitle(title)
            if windows:
                window = windows[0]
                window.close()
                context.log(f"已關閉視窗: {window.title}")
                return {"closed": True, "title": window.title}
            else:
                context.log(f"未找到視窗: {title}")
                return {"closed": False}
        except Exception as e:
            context.log(f"關閉視窗失敗: {e}", level="warning")
            return {"closed": False, "error": str(e)}


class ResizeWindowExecutor(BlockExecutor):
    """調整視窗大小"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import pyautogui
        
        title = context.resolve_value(block.params.get("title", ""))
        width = int(context.resolve_value(block.params.get("width", 800)))
        height = int(context.resolve_value(block.params.get("height", 600)))
        
        try:
            windows = pyautogui.getWindowsWithTitle(title)
            if windows:
                window = windows[0]
                window.resizeTo(width, height)
                context.log(f"已調整視窗大小: {width}x{height}")
                return {"resized": True, "width": width, "height": height}
            else:
                context.log(f"未找到視窗: {title}")
                return {"resized": False}
        except Exception as e:
            context.log(f"調整視窗大小失敗: {e}", level="warning")
            return {"resized": False, "error": str(e)}


class MoveWindowExecutor(BlockExecutor):
    """移動視窗"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        import pyautogui
        
        title = context.resolve_value(block.params.get("title", ""))
        x = int(context.resolve_value(block.params.get("x", 0)))
        y = int(context.resolve_value(block.params.get("y", 0)))
        
        try:
            windows = pyautogui.getWindowsWithTitle(title)
            if windows:
                window = windows[0]
                window.moveTo(x, y)
                context.log(f"已移動視窗到: ({x}, {y})")
                return {"moved": True, "x": x, "y": y}
            else:
                context.log(f"未找到視窗: {title}")
                return {"moved": False}
        except Exception as e:
            context.log(f"移動視窗失敗: {e}", level="warning")
            return {"moved": False, "error": str(e)}


# ==================== OCR Blocks ====================

class OcrReadTextExecutor(BlockExecutor):
    """OCR 讀取文字"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        try:
            import pytesseract
            from PIL import Image
            import io
        except ImportError:
            context.log("OCR 功能需要安裝 pytesseract", level="error")
            return {"text": "", "success": False, "error": "pytesseract not installed"}
        
        region = block.params.get("region")  # (x, y, width, height)
        language = block.params.get("language", "eng+chi_tra")
        
        context.log(f"執行 OCR 識別 (語言: {language})")
        
        try:
            # 擷取螢幕區域
            if region:
                x, y, w, h = region
                screenshot_bytes = image_finder.capture_region(x, y, w, h)
            else:
                screenshot_bytes = image_finder.capture_screen()
            
            # 轉換為 PIL Image
            image = Image.open(io.BytesIO(screenshot_bytes))
            
            # 執行 OCR
            text = pytesseract.image_to_string(image, lang=language)
            text = text.strip()
            
            context.log(f"OCR 識別結果: {text[:100]}{'...' if len(text) > 100 else ''}")
            
            return {"text": text, "success": True}
        except Exception as e:
            context.log(f"OCR 識別失敗: {e}", level="error")
            return {"text": "", "success": False, "error": str(e)}


class OcrFindTextExecutor(BlockExecutor):
    """OCR 尋找文字位置"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        try:
            import pytesseract
            from PIL import Image
            import io
        except ImportError:
            context.log("OCR 功能需要安裝 pytesseract", level="error")
            return {"found": False, "error": "pytesseract not installed"}
        
        search_text = context.resolve_value(block.params.get("text", ""))
        language = block.params.get("language", "eng+chi_tra")
        
        context.log(f"OCR 尋找文字: {search_text}")
        
        try:
            # 擷取全螢幕
            screenshot_bytes = image_finder.capture_screen()
            image = Image.open(io.BytesIO(screenshot_bytes))
            
            # 取得文字位置資訊
            data = pytesseract.image_to_data(image, lang=language, output_type=pytesseract.Output.DICT)
            
            # 搜尋文字
            for i, text in enumerate(data['text']):
                if search_text.lower() in text.lower():
                    x = data['left'][i]
                    y = data['top'][i]
                    w = data['width'][i]
                    h = data['height'][i]
                    center_x = x + w // 2
                    center_y = y + h // 2
                    
                    context.log(f"找到文字 '{search_text}' 於 ({center_x}, {center_y})")
                    
                    return {
                        "found": True,
                        "x": center_x,
                        "y": center_y,
                        "width": w,
                        "height": h,
                        "text": text,
                    }
            
            context.log(f"未找到文字: {search_text}")
            return {"found": False}
            
        except Exception as e:
            context.log(f"OCR 尋找失敗: {e}", level="error")
            return {"found": False, "error": str(e)}


class OcrClickTextExecutor(BlockExecutor):
    """OCR 點擊文字"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        try:
            import pytesseract
            from PIL import Image
            import io
        except ImportError:
            context.log("OCR 功能需要安裝 pytesseract", level="error")
            return {"clicked": False, "error": "pytesseract not installed"}
        
        search_text = context.resolve_value(block.params.get("text", ""))
        language = block.params.get("language", "eng+chi_tra")
        timeout = block.params.get("timeout", 30)
        
        context.log(f"OCR 點擊文字: {search_text}")
        
        import time
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            if context.should_stop:
                return {"clicked": False, "cancelled": True}
            
            try:
                screenshot_bytes = image_finder.capture_screen()
                image = Image.open(io.BytesIO(screenshot_bytes))
                
                data = pytesseract.image_to_data(image, lang=language, output_type=pytesseract.Output.DICT)
                
                for i, text in enumerate(data['text']):
                    if search_text.lower() in text.lower():
                        x = data['left'][i]
                        y = data['top'][i]
                        w = data['width'][i]
                        h = data['height'][i]
                        center_x = x + w // 2
                        center_y = y + h // 2
                        
                        context.log(f"找到並點擊文字 '{search_text}' 於 ({center_x}, {center_y})")
                        desktop.click(center_x, center_y)
                        
                        return {"clicked": True, "x": center_x, "y": center_y}
                        
            except Exception as e:
                context.log(f"OCR 搜尋錯誤: {e}", level="warning")
            
            await asyncio.sleep(0.5)
        
        context.log(f"OCR 點擊超時: {search_text}")
        return {"clicked": False, "timeout": True}


class OcrWaitTextExecutor(BlockExecutor):
    """OCR 等待文字出現"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        try:
            import pytesseract
            from PIL import Image
            import io
        except ImportError:
            context.log("OCR 功能需要安裝 pytesseract", level="error")
            return {"found": False, "error": "pytesseract not installed"}
        
        search_text = context.resolve_value(block.params.get("text", ""))
        language = block.params.get("language", "eng+chi_tra")
        timeout = block.params.get("timeout", 30)
        
        context.log(f"OCR 等待文字出現: {search_text}")
        
        import time
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            if context.should_stop:
                return {"found": False, "cancelled": True}
            
            try:
                screenshot_bytes = image_finder.capture_screen()
                image = Image.open(io.BytesIO(screenshot_bytes))
                
                text = pytesseract.image_to_string(image, lang=language)
                
                if search_text.lower() in text.lower():
                    context.log(f"文字已出現: {search_text}")
                    return {"found": True, "text": text}
                    
            except Exception as e:
                context.log(f"OCR 搜尋錯誤: {e}", level="warning")
            
            await asyncio.sleep(0.5)
        
        context.log(f"OCR 等待超時: {search_text}")
        return {"found": False, "timeout": True}


class IfOcrTextExistsExecutor(BlockExecutor):
    """如果 OCR 文字存在"""
    
    async def execute(self, block: BlockInstance, context: ExecutionContext) -> Any:
        try:
            import pytesseract
            from PIL import Image
            import io
        except ImportError:
            context.log("OCR 功能需要安裝 pytesseract", level="error")
            return {"exists": False, "execute_else": True}
        
        search_text = context.resolve_value(block.params.get("text", ""))
        language = block.params.get("language", "eng+chi_tra")
        
        context.log(f"OCR 檢查文字是否存在: {search_text}")
        
        try:
            screenshot_bytes = image_finder.capture_screen()
            image = Image.open(io.BytesIO(screenshot_bytes))
            
            text = pytesseract.image_to_string(image, lang=language)
            exists = search_text.lower() in text.lower()
            
            if exists:
                context.log("文字存在，執行子區塊")
                return {"exists": True, "execute_children": True}
            else:
                context.log("文字不存在，執行 else 區塊")
                return {"exists": False, "execute_else": True}
                
        except Exception as e:
            context.log(f"OCR 檢查失敗: {e}", level="error")
            return {"exists": False, "execute_else": True, "error": str(e)}


# ==================== Block 註冊表 ====================

class BlockRegistry:
    """Block 註冊表"""
    
    # Block 定義
    DEFINITIONS: Dict[str, BlockDefinition] = {
        # 動作類
        "click_image": BlockDefinition(
            id="click_image",
            name="點擊圖片",
            category=BlockCategory.ACTION,
            description="找到圖片並點擊",
            params={
                "image_path": {"type": "image", "required": True, "description": "圖片路徑"},
                "confidence": {"type": "number", "default": 0.8, "min": 0.1, "max": 1.0, "step": 0.05, "description": "匹配信心度 (0.1-1.0)"},
                "timeout": {"type": "number", "default": 30, "description": "超時時間（秒）"},
                "offset_x": {"type": "number", "default": 0, "description": "X 偏移"},
                "offset_y": {"type": "number", "default": 0, "description": "Y 偏移"},
            },
            color="#5b80a5",
        ),
        "click_position": BlockDefinition(
            id="click_position",
            name="點擊座標",
            category=BlockCategory.ACTION,
            description="點擊指定座標",
            params={
                "x": {"type": "number", "required": True, "description": "X 座標"},
                "y": {"type": "number", "required": True, "description": "Y 座標"},
                "button": {"type": "select", "options": ["left", "right", "middle"], "default": "left", "description": "滑鼠按鈕"},
            },
            color="#5b80a5",
        ),
        "double_click_image": BlockDefinition(
            id="double_click_image",
            name="雙擊圖片",
            category=BlockCategory.ACTION,
            description="找到圖片並雙擊",
            params={
                "image_path": {"type": "image", "required": True, "description": "圖片路徑"},
                "confidence": {"type": "number", "default": 0.8, "min": 0.1, "max": 1.0, "step": 0.05, "description": "匹配信心度 (0.1-1.0)"},
                "timeout": {"type": "number", "default": 30, "description": "超時時間（秒）"},
            },
            color="#5b80a5",
        ),
        "right_click_image": BlockDefinition(
            id="right_click_image",
            name="右鍵圖片",
            category=BlockCategory.ACTION,
            description="找到圖片並右鍵點擊",
            params={
                "image_path": {"type": "image", "required": True, "description": "圖片路徑"},
                "confidence": {"type": "number", "default": 0.8, "min": 0.1, "max": 1.0, "step": 0.05, "description": "匹配信心度 (0.1-1.0)"},
                "timeout": {"type": "number", "default": 30, "description": "超時時間（秒）"},
            },
            color="#5b80a5",
        ),
        "type_text": BlockDefinition(
            id="type_text",
            name="輸入文字",
            category=BlockCategory.ACTION,
            description="模擬鍵盤輸入文字",
            params={
                "text": {"type": "text", "required": True, "description": "要輸入的文字"},
                "interval": {"type": "number", "default": 0.05, "description": "字元間隔（秒）"},
            },
            color="#5b80a5",
        ),
        "hotkey": BlockDefinition(
            id="hotkey",
            name="快捷鍵",
            category=BlockCategory.ACTION,
            description="執行組合鍵",
            params={
                "keys": {"type": "keys", "required": True, "description": "按鍵組合（如 ctrl+c）"},
            },
            color="#5b80a5",
        ),
        "scroll": BlockDefinition(
            id="scroll",
            name="滾動",
            category=BlockCategory.ACTION,
            description="滾動滑鼠滾輪",
            params={
                "direction": {"type": "select", "options": ["up", "down"], "default": "down", "description": "方向"},
                "amount": {"type": "number", "default": 3, "description": "滾動量"},
                "x": {"type": "number", "description": "X 座標（可選）"},
                "y": {"type": "number", "description": "Y 座標（可選）"},
            },
            color="#5b80a5",
        ),
        "drag_drop": BlockDefinition(
            id="drag_drop",
            name="拖放",
            category=BlockCategory.ACTION,
            description="從 A 圖片拖到 B 圖片",
            params={
                "from_image": {"type": "image", "required": True, "description": "起點圖片"},
                "to_image": {"type": "image", "required": True, "description": "終點圖片"},
                "confidence": {"type": "number", "default": 0.8, "min": 0.1, "max": 1.0, "step": 0.05, "description": "匹配信心度 (0.1-1.0)"},
                "timeout": {"type": "number", "default": 30, "description": "超時時間（秒）"},
            },
            color="#5b80a5",
        ),
        "wait": BlockDefinition(
            id="wait",
            name="等待",
            category=BlockCategory.ACTION,
            description="等待指定秒數",
            params={
                "seconds": {"type": "number", "required": True, "default": 1, "description": "等待秒數"},
            },
            color="#5b80a5",
        ),
        "wait_image": BlockDefinition(
            id="wait_image",
            name="等待圖片出現",
            category=BlockCategory.ACTION,
            description="等待直到圖片出現",
            params={
                "image_path": {"type": "image", "required": True, "description": "圖片路徑"},
                "confidence": {"type": "number", "default": 0.8, "min": 0.1, "max": 1.0, "step": 0.05, "description": "匹配信心度 (0.1-1.0)"},
                "timeout": {"type": "number", "default": 30, "description": "超時時間（秒）"},
            },
            color="#5b80a5",
        ),
        "wait_image_gone": BlockDefinition(
            id="wait_image_gone",
            name="等待圖片消失",
            category=BlockCategory.ACTION,
            description="等待直到圖片消失",
            params={
                "image_path": {"type": "image", "required": True, "description": "圖片路徑"},
                "confidence": {"type": "number", "default": 0.8, "min": 0.1, "max": 1.0, "step": 0.05, "description": "匹配信心度 (0.1-1.0)"},
                "timeout": {"type": "number", "default": 30, "description": "超時時間（秒）"},
            },
            color="#5b80a5",
        ),
        
        # 控制類
        "if_image_exists": BlockDefinition(
            id="if_image_exists",
            name="如果圖片存在",
            category=BlockCategory.CONTROL,
            description="條件分支：如果圖片存在",
            params={
                "image_path": {"type": "image", "required": True, "description": "圖片路徑"},
                "confidence": {"type": "number", "default": 0.8, "min": 0.1, "max": 1.0, "step": 0.05, "description": "匹配信心度 (0.1-1.0)"},
            },
            has_children=True,
            has_else=True,
            color="#5ca65b",
        ),
        "loop_times": BlockDefinition(
            id="loop_times",
            name="重複 N 次",
            category=BlockCategory.CONTROL,
            description="固定次數迴圈",
            params={
                "times": {"type": "number", "required": True, "default": 1, "description": "重複次數"},
            },
            has_children=True,
            color="#5ca65b",
        ),
        "loop_while_image": BlockDefinition(
            id="loop_while_image",
            name="當圖片存在時重複",
            category=BlockCategory.CONTROL,
            description="條件迴圈：當圖片存在時重複",
            params={
                "image_path": {"type": "image", "required": True, "description": "圖片路徑"},
                "confidence": {"type": "number", "default": 0.8, "min": 0.1, "max": 1.0, "step": 0.05, "description": "匹配信心度 (0.1-1.0)"},
                "max_iterations": {"type": "number", "default": 100, "description": "最大迭代次數"},
            },
            has_children=True,
            color="#5ca65b",
        ),
        "loop_until_image": BlockDefinition(
            id="loop_until_image",
            name="重複直到圖片出現",
            category=BlockCategory.CONTROL,
            description="條件迴圈：重複直到圖片出現",
            params={
                "image_path": {"type": "image", "required": True, "description": "圖片路徑"},
                "confidence": {"type": "number", "default": 0.8, "min": 0.1, "max": 1.0, "step": 0.05, "description": "匹配信心度 (0.1-1.0)"},
                "max_iterations": {"type": "number", "default": 100, "description": "最大迭代次數"},
            },
            has_children=True,
            color="#5ca65b",
        ),
        "break": BlockDefinition(
            id="break",
            name="跳出迴圈",
            category=BlockCategory.CONTROL,
            description="中斷當前迴圈",
            params={},
            color="#5ca65b",
        ),
        "continue": BlockDefinition(
            id="continue",
            name="繼續下一輪",
            category=BlockCategory.CONTROL,
            description="跳過本輪迴圈",
            params={},
            color="#5ca65b",
        ),
        
        # 變數類
        "set_variable": BlockDefinition(
            id="set_variable",
            name="設定變數",
            category=BlockCategory.VARIABLE,
            description="設定變數值",
            params={
                "name": {"type": "text", "required": True, "description": "變數名稱"},
                "value": {"type": "any", "required": True, "description": "變數值"},
            },
            color="#a55b80",
        ),
        "get_variable": BlockDefinition(
            id="get_variable",
            name="取得變數",
            category=BlockCategory.VARIABLE,
            description="讀取變數值",
            params={
                "name": {"type": "text", "required": True, "description": "變數名稱"},
            },
            color="#a55b80",
        ),
        "save_position": BlockDefinition(
            id="save_position",
            name="儲存座標",
            category=BlockCategory.VARIABLE,
            description="將找到的座標存入變數",
            params={
                "variable_name": {"type": "text", "required": True, "description": "變數名稱"},
                "image_path": {"type": "image", "required": True, "description": "圖片路徑"},
                "confidence": {"type": "number", "default": 0.8, "min": 0.1, "max": 1.0, "step": 0.05, "description": "匹配信心度 (0.1-1.0)"},
                "timeout": {"type": "number", "default": 30, "description": "超時時間（秒）"},
            },
            color="#a55b80",
        ),
        
        # 進階類
        "run_script": BlockDefinition(
            id="run_script",
            name="執行子腳本",
            category=BlockCategory.ADVANCED,
            description="呼叫另一個腳本",
            params={
                "script_id": {"type": "script", "required": True, "description": "腳本 ID"},
            },
            color="#a5805b",
        ),
        "http_request": BlockDefinition(
            id="http_request",
            name="HTTP 請求",
            category=BlockCategory.ADVANCED,
            description="發送 HTTP 請求",
            params={
                "method": {"type": "select", "options": ["GET", "POST", "PUT", "DELETE", "PATCH"], "default": "GET", "description": "HTTP 方法"},
                "url": {"type": "text", "required": True, "description": "URL"},
                "headers": {"type": "json", "default": {}, "description": "請求標頭"},
                "body": {"type": "text", "description": "請求內容"},
            },
            color="#a5805b",
        ),
        "run_command": BlockDefinition(
            id="run_command",
            name="執行命令",
            category=BlockCategory.ADVANCED,
            description="執行系統命令",
            params={
                "command": {"type": "text", "required": True, "description": "命令"},
            },
            color="#a5805b",
        ),
        "log": BlockDefinition(
            id="log",
            name="記錄日誌",
            category=BlockCategory.ADVANCED,
            description="輸出日誌",
            params={
                "message": {"type": "text", "required": True, "description": "日誌訊息"},
                "level": {"type": "select", "options": ["debug", "info", "warning", "error"], "default": "info", "description": "日誌等級"},
            },
            color="#a5805b",
        ),
        "screenshot": BlockDefinition(
            id="screenshot",
            name="截圖",
            category=BlockCategory.ADVANCED,
            description="儲存當前螢幕截圖",
            params={
                "save_path": {"type": "text", "description": "儲存路徑（可選）"},
            },
            color="#a5805b",
        ),
        "send_email": BlockDefinition(
            id="send_email",
            name="發送郵件",
            category=BlockCategory.ADVANCED,
            description="發送 Email 通知",
            params={
                "to": {"type": "text", "required": True, "description": "收件人"},
                "subject": {"type": "text", "required": True, "description": "主旨"},
                "body": {"type": "text", "required": True, "description": "內容"},
            },
            color="#EC4899",
        ),
        "retry": BlockDefinition(
            id="retry",
            name="重試",
            category=BlockCategory.CONTROL,
            description="失敗時自動重試",
            params={
                "times": {"type": "number", "default": 3, "description": "重試次數"},
                "interval": {"type": "number", "default": 1, "description": "間隔秒數"},
            },
            color="#8B5CF6",
            has_children=True,
        ),
        
        # 檔案操作
        "read_file": BlockDefinition(
            id="read_file",
            name="讀取檔案",
            category=BlockCategory.ADVANCED,
            description="讀取檔案內容",
            params={
                "file_path": {"type": "text", "required": True, "description": "檔案路徑"},
                "encoding": {"type": "text", "default": "utf-8", "description": "編碼"},
            },
            color="#a5805b",
        ),
        "write_file": BlockDefinition(
            id="write_file",
            name="寫入檔案",
            category=BlockCategory.ADVANCED,
            description="寫入檔案內容",
            params={
                "file_path": {"type": "text", "required": True, "description": "檔案路徑"},
                "content": {"type": "text", "required": True, "description": "內容"},
                "encoding": {"type": "text", "default": "utf-8", "description": "編碼"},
                "append": {"type": "boolean", "default": False, "description": "附加模式"},
            },
            color="#a5805b",
        ),
        "copy_file": BlockDefinition(
            id="copy_file",
            name="複製檔案",
            category=BlockCategory.ADVANCED,
            description="複製檔案或資料夾",
            params={
                "source": {"type": "text", "required": True, "description": "來源路徑"},
                "destination": {"type": "text", "required": True, "description": "目的路徑"},
            },
            color="#a5805b",
        ),
        "move_file": BlockDefinition(
            id="move_file",
            name="移動檔案",
            category=BlockCategory.ADVANCED,
            description="移動檔案或資料夾",
            params={
                "source": {"type": "text", "required": True, "description": "來源路徑"},
                "destination": {"type": "text", "required": True, "description": "目的路徑"},
            },
            color="#a5805b",
        ),
        "delete_file": BlockDefinition(
            id="delete_file",
            name="刪除檔案",
            category=BlockCategory.ADVANCED,
            description="刪除檔案或資料夾",
            params={
                "file_path": {"type": "text", "required": True, "description": "檔案路徑"},
            },
            color="#a5805b",
        ),
        "file_exists": BlockDefinition(
            id="file_exists",
            name="檔案存在",
            category=BlockCategory.CONTROL,
            description="檢查檔案是否存在",
            params={
                "file_path": {"type": "text", "required": True, "description": "檔案路徑"},
            },
            has_children=True,
            has_else=True,
            color="#5ca65b",
        ),
        "list_files": BlockDefinition(
            id="list_files",
            name="列出檔案",
            category=BlockCategory.ADVANCED,
            description="列出目錄下的檔案",
            params={
                "directory": {"type": "text", "required": True, "description": "目錄路徑"},
                "pattern": {"type": "text", "default": "*", "description": "匹配模式"},
            },
            color="#a5805b",
        ),
        
        # Excel 操作
        "excel_read": BlockDefinition(
            id="excel_read",
            name="讀取 Excel",
            category=BlockCategory.ADVANCED,
            description="讀取 Excel 檔案",
            params={
                "file_path": {"type": "text", "required": True, "description": "檔案路徑"},
                "sheet": {"type": "text", "default": "Sheet1", "description": "工作表名稱"},
                "var_name": {"type": "text", "default": "data", "description": "變數名稱"},
            },
            color="#10B981",
        ),
        "excel_write": BlockDefinition(
            id="excel_write",
            name="寫入 Excel",
            category=BlockCategory.ADVANCED,
            description="寫入 Excel 檔案",
            params={
                "file_path": {"type": "text", "required": True, "description": "檔案路徑"},
                "sheet": {"type": "text", "default": "Sheet1", "description": "工作表名稱"},
                "data_var": {"type": "text", "required": True, "description": "資料變數"},
            },
            color="#10B981",
        ),
        "excel_read_cell": BlockDefinition(
            id="excel_read_cell",
            name="讀取 Excel 儲存格",
            category=BlockCategory.ADVANCED,
            description="讀取 Excel 指定儲存格",
            params={
                "file_path": {"type": "text", "required": True, "description": "檔案路徑"},
                "sheet": {"type": "text", "default": "Sheet1", "description": "工作表名稱"},
                "cell": {"type": "text", "required": True, "description": "儲存格 (如 A1)"},
                "var_name": {"type": "text", "default": "value", "description": "變數名稱"},
            },
            color="#10B981",
        ),
        "excel_write_cell": BlockDefinition(
            id="excel_write_cell",
            name="寫入 Excel 儲存格",
            category=BlockCategory.ADVANCED,
            description="寫入 Excel 指定儲存格",
            params={
                "file_path": {"type": "text", "required": True, "description": "檔案路徑"},
                "sheet": {"type": "text", "default": "Sheet1", "description": "工作表名稱"},
                "cell": {"type": "text", "required": True, "description": "儲存格 (如 A1)"},
                "value": {"type": "text", "required": True, "description": "值"},
            },
            color="#10B981",
        ),
        
        # 錯誤處理
        "try_catch": BlockDefinition(
            id="try_catch",
            name="錯誤處理",
            category=BlockCategory.CONTROL,
            description="Try-Catch 錯誤處理",
            params={},
            has_children=True,
            has_else=True,
            color="#5ca65b",
        ),
        
        # 迴圈增強
        "for_each": BlockDefinition(
            id="for_each",
            name="遍歷列表",
            category=BlockCategory.CONTROL,
            description="遍歷列表中的每個項目",
            params={
                "list": {"type": "text", "required": True, "description": "列表（JSON 格式或逗號分隔）"},
                "variable_name": {"type": "text", "default": "item", "description": "項目變數名稱"},
            },
            has_children=True,
            color="#5ca65b",
        ),
        
        # 字串處理
        "string_concat": BlockDefinition(
            id="string_concat",
            name="字串拼接",
            category=BlockCategory.VARIABLE,
            description="拼接兩個字串",
            params={
                "string1": {"type": "text", "required": True, "description": "字串1"},
                "string2": {"type": "text", "required": True, "description": "字串2"},
                "separator": {"type": "text", "default": "", "description": "分隔符"},
            },
            color="#a55b80",
        ),
        "string_split": BlockDefinition(
            id="string_split",
            name="字串分割",
            category=BlockCategory.VARIABLE,
            description="分割字串為列表",
            params={
                "text": {"type": "text", "required": True, "description": "文字"},
                "separator": {"type": "text", "default": ",", "description": "分隔符"},
            },
            color="#a55b80",
        ),
        "string_replace": BlockDefinition(
            id="string_replace",
            name="字串替換",
            category=BlockCategory.VARIABLE,
            description="替換字串內容",
            params={
                "text": {"type": "text", "required": True, "description": "文字"},
                "search": {"type": "text", "required": True, "description": "搜尋"},
                "replace": {"type": "text", "default": "", "description": "替換為"},
                "use_regex": {"type": "boolean", "default": False, "description": "使用正則"},
            },
            color="#a55b80",
        ),
        "string_match": BlockDefinition(
            id="string_match",
            name="正則匹配",
            category=BlockCategory.VARIABLE,
            description="正則表達式匹配",
            params={
                "text": {"type": "text", "required": True, "description": "文字"},
                "pattern": {"type": "text", "required": True, "description": "正則模式"},
            },
            color="#a55b80",
        ),
        "string_format": BlockDefinition(
            id="string_format",
            name="字串格式化",
            category=BlockCategory.VARIABLE,
            description="格式化字串模板",
            params={
                "template": {"type": "text", "required": True, "description": "模板（使用 ${變數名}）"},
            },
            color="#a55b80",
        ),
        
        # 數學運算
        "math_operation": BlockDefinition(
            id="math_operation",
            name="數學運算",
            category=BlockCategory.VARIABLE,
            description="執行數學運算",
            params={
                "a": {"type": "number", "required": True, "description": "數值 A"},
                "operation": {"type": "select", "options": ["add", "subtract", "multiply", "divide", "modulo", "power"], "default": "add", "description": "運算"},
                "b": {"type": "number", "required": True, "description": "數值 B"},
            },
            color="#a55b80",
        ),
        "random_number": BlockDefinition(
            id="random_number",
            name="隨機數",
            category=BlockCategory.VARIABLE,
            description="生成隨機數",
            params={
                "min": {"type": "number", "default": 0, "description": "最小值"},
                "max": {"type": "number", "default": 100, "description": "最大值"},
                "integer": {"type": "boolean", "default": True, "description": "整數"},
            },
            color="#a55b80",
        ),
        
        # 剪貼簿
        "clipboard_read": BlockDefinition(
            id="clipboard_read",
            name="讀取剪貼簿",
            category=BlockCategory.VARIABLE,
            description="讀取剪貼簿內容",
            params={},
            color="#a55b80",
        ),
        "clipboard_write": BlockDefinition(
            id="clipboard_write",
            name="寫入剪貼簿",
            category=BlockCategory.VARIABLE,
            description="寫入內容到剪貼簿",
            params={
                "content": {"type": "text", "required": True, "description": "內容"},
            },
            color="#a55b80",
        ),
        
        # JSON 操作
        "json_parse": BlockDefinition(
            id="json_parse",
            name="解析 JSON",
            category=BlockCategory.VARIABLE,
            description="解析 JSON 字串",
            params={
                "text": {"type": "text", "required": True, "description": "JSON 字串"},
            },
            color="#a55b80",
        ),
        "json_stringify": BlockDefinition(
            id="json_stringify",
            name="轉換 JSON",
            category=BlockCategory.VARIABLE,
            description="轉換為 JSON 字串",
            params={
                "data": {"type": "any", "required": True, "description": "資料"},
                "pretty": {"type": "boolean", "default": False, "description": "美化輸出"},
            },
            color="#a55b80",
        ),
        "json_get_value": BlockDefinition(
            id="json_get_value",
            name="取得 JSON 值",
            category=BlockCategory.VARIABLE,
            description="從 JSON 取得值",
            params={
                "data": {"type": "any", "required": True, "description": "JSON 資料"},
                "path": {"type": "text", "required": True, "description": "路徑（如 user.name）"},
            },
            color="#a55b80",
        ),
        
        # 視窗控制
        "get_window": BlockDefinition(
            id="get_window",
            name="取得視窗",
            category=BlockCategory.ADVANCED,
            description="取得視窗資訊",
            params={
                "title": {"type": "text", "required": True, "description": "視窗標題"},
            },
            color="#a5805b",
        ),
        "activate_window": BlockDefinition(
            id="activate_window",
            name="啟用視窗",
            category=BlockCategory.ADVANCED,
            description="啟用（聚焦）視窗",
            params={
                "title": {"type": "text", "required": True, "description": "視窗標題"},
            },
            color="#a5805b",
        ),
        "close_window": BlockDefinition(
            id="close_window",
            name="關閉視窗",
            category=BlockCategory.ADVANCED,
            description="關閉視窗",
            params={
                "title": {"type": "text", "required": True, "description": "視窗標題"},
            },
            color="#a5805b",
        ),
        "resize_window": BlockDefinition(
            id="resize_window",
            name="調整視窗大小",
            category=BlockCategory.ADVANCED,
            description="調整視窗大小",
            params={
                "title": {"type": "text", "required": True, "description": "視窗標題"},
                "width": {"type": "number", "required": True, "description": "寬度"},
                "height": {"type": "number", "required": True, "description": "高度"},
            },
            color="#a5805b",
        ),
        "move_window": BlockDefinition(
            id="move_window",
            name="移動視窗",
            category=BlockCategory.ADVANCED,
            description="移動視窗位置",
            params={
                "title": {"type": "text", "required": True, "description": "視窗標題"},
                "x": {"type": "number", "required": True, "description": "X 座標"},
                "y": {"type": "number", "required": True, "description": "Y 座標"},
            },
            color="#a5805b",
        ),
        
        # OCR
        "ocr_read_text": BlockDefinition(
            id="ocr_read_text",
            name="OCR 讀取文字",
            category=BlockCategory.ACTION,
            description="使用 OCR 讀取螢幕上的文字",
            params={
                "language": {"type": "text", "default": "eng+chi_tra", "description": "語言 (如 eng, chi_tra)"},
            },
            color="#5b80a5",
        ),
        "ocr_find_text": BlockDefinition(
            id="ocr_find_text",
            name="OCR 尋找文字",
            category=BlockCategory.ACTION,
            description="使用 OCR 尋找文字位置",
            params={
                "text": {"type": "text", "required": True, "description": "要尋找的文字"},
                "language": {"type": "text", "default": "eng+chi_tra", "description": "語言"},
            },
            color="#5b80a5",
        ),
        "ocr_click_text": BlockDefinition(
            id="ocr_click_text",
            name="OCR 點擊文字",
            category=BlockCategory.ACTION,
            description="使用 OCR 尋找並點擊文字",
            params={
                "text": {"type": "text", "required": True, "description": "要點擊的文字"},
                "language": {"type": "text", "default": "eng+chi_tra", "description": "語言"},
                "timeout": {"type": "number", "default": 30, "description": "超時時間（秒）"},
            },
            color="#5b80a5",
        ),
        "ocr_wait_text": BlockDefinition(
            id="ocr_wait_text",
            name="OCR 等待文字",
            category=BlockCategory.ACTION,
            description="等待直到文字出現",
            params={
                "text": {"type": "text", "required": True, "description": "要等待的文字"},
                "language": {"type": "text", "default": "eng+chi_tra", "description": "語言"},
                "timeout": {"type": "number", "default": 30, "description": "超時時間（秒）"},
            },
            color="#5b80a5",
        ),
        "if_ocr_text_exists": BlockDefinition(
            id="if_ocr_text_exists",
            name="如果 OCR 文字存在",
            category=BlockCategory.CONTROL,
            description="使用 OCR 檢查文字是否存在",
            params={
                "text": {"type": "text", "required": True, "description": "要檢查的文字"},
                "language": {"type": "text", "default": "eng+chi_tra", "description": "語言"},
            },
            has_children=True,
            has_else=True,
            color="#5ca65b",
        ),
    }
    
    # Block 執行器映射
    EXECUTORS: Dict[str, Type[BlockExecutor]] = {
        # 動作類
        "click_image": ClickImageExecutor,
        "click_position": ClickPositionExecutor,
        "double_click_image": DoubleClickImageExecutor,
        "right_click_image": RightClickImageExecutor,
        "type_text": TypeTextExecutor,
        "hotkey": HotkeyExecutor,
        "scroll": ScrollExecutor,
        "drag_drop": DragDropExecutor,
        "wait": WaitExecutor,
        "wait_image": WaitImageExecutor,
        "wait_image_gone": WaitImageGoneExecutor,
        
        # 控制類
        "if_image_exists": IfImageExistsExecutor,
        "loop_times": LoopTimesExecutor,
        "loop_while_image": LoopWhileImageExecutor,
        "loop_until_image": LoopUntilImageExecutor,
        "break": BreakExecutor,
        "continue": ContinueExecutor,
        "file_exists": FileExistsExecutor,
        "try_catch": TryCatchExecutor,
        "for_each": ForEachExecutor,
        
        # 變數類
        "set_variable": SetVariableExecutor,
        "get_variable": GetVariableExecutor,
        "save_position": SavePositionExecutor,
        "string_concat": StringConcatExecutor,
        "string_split": StringSplitExecutor,
        "string_replace": StringReplaceExecutor,
        "string_match": StringMatchExecutor,
        "string_format": StringFormatExecutor,
        "math_operation": MathOperationExecutor,
        "random_number": RandomNumberExecutor,
        "clipboard_read": ClipboardReadExecutor,
        "clipboard_write": ClipboardWriteExecutor,
        "json_parse": JsonParseExecutor,
        "json_stringify": JsonStringifyExecutor,
        "json_get_value": JsonGetValueExecutor,
        
        # 進階類
        "run_script": RunScriptExecutor,
        "http_request": HttpRequestExecutor,
        "run_command": RunCommandExecutor,
        "log": LogExecutor,
        "screenshot": ScreenshotExecutor,
        "send_email": SendEmailExecutor,
        "retry": RetryExecutor,
        "read_file": ReadFileExecutor,
        "write_file": WriteFileExecutor,
        "copy_file": CopyFileExecutor,
        "move_file": MoveFileExecutor,
        "delete_file": DeleteFileExecutor,
        "list_files": ListFilesExecutor,
        
        # Excel
        "excel_read": ExcelReadExecutor,
        "excel_write": ExcelWriteExecutor,
        "excel_read_cell": ExcelReadCellExecutor,
        "excel_write_cell": ExcelWriteCellExecutor,
        
        "get_window": GetWindowExecutor,
        "activate_window": ActivateWindowExecutor,
        "close_window": CloseWindowExecutor,
        "resize_window": ResizeWindowExecutor,
        "move_window": MoveWindowExecutor,
        
        # OCR
        "ocr_read_text": OcrReadTextExecutor,
        "ocr_find_text": OcrFindTextExecutor,
        "ocr_click_text": OcrClickTextExecutor,
        "ocr_wait_text": OcrWaitTextExecutor,
        "if_ocr_text_exists": IfOcrTextExistsExecutor,
    }
    
    @classmethod
    def get_definition(cls, block_id: str) -> Optional[BlockDefinition]:
        """取得 Block 定義"""
        return cls.DEFINITIONS.get(block_id)
    
    @classmethod
    def get_executor(cls, block_id: str) -> Optional[BlockExecutor]:
        """取得 Block 執行器實例"""
        executor_class = cls.EXECUTORS.get(block_id)
        if executor_class:
            return executor_class()
        return None
    
    @classmethod
    def get_all_definitions(cls) -> Dict[str, BlockDefinition]:
        """取得所有 Block 定義"""
        return cls.DEFINITIONS
    
    @classmethod
    def get_definitions_by_category(cls, category: BlockCategory) -> Dict[str, BlockDefinition]:
        """依類別取得 Block 定義"""
        return {
            k: v for k, v in cls.DEFINITIONS.items()
            if v.category == category
        }
